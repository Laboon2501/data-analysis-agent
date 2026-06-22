"""文件数据源导入工具，将 CSV/Excel/Parquet 转换为只读 SQLite 表。"""

from __future__ import annotations

import csv
import re
import sqlite3
from collections.abc import Iterable
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from schemas._base import StrictBaseModel, utc_now

ALLOWED_FILE_SUFFIXES = frozenset({".csv", ".xlsx", ".parquet"})
SENSITIVE_FILE_NAMES = frozenset({".env", ".env.local", ".env.dev", ".env.prod"})
SENSITIVE_PATH_PARTS = frozenset({".ssh", ".aws", ".azure", ".gcp", "system32", "etc"})
IDENTIFIER_PATTERN = re.compile(r"[^A-Za-z0-9_]+")


class FileDataSourceImportResult(StrictBaseModel):
    """导入文件后的安全元数据和内部 SQLite 路径。"""

    sqlite_path: str
    original_filename: str
    source_type: str
    file_size: int
    table_name: str
    uploaded_at: str
    row_count: int
    columns: list[str]


def import_file_to_sqlite(
    *,
    source_path: Path,
    datasource_id: str,
    output_dir: Path,
    table_name: str | None = None,
    original_filename: str | None = None,
    source_type: str = "path",
    max_bytes: int | None = None,
) -> FileDataSourceImportResult:
    """读取受支持文件并导入到受控 SQLite 数据库文件。

    该函数只生成一个普通 SQLite 文件，后续所有查询仍走 SQLAlchemyDataSource
    和 SQLGuard；不会把原始文件内容写入 events/history。
    """

    safe_source = validate_supported_file_path(source_path, max_bytes=max_bytes)
    output_dir.mkdir(parents=True, exist_ok=True)
    rows = _read_rows(safe_source)
    if not rows or not _has_value(rows[0]):
        raise ValueError("File datasource is empty or missing a header row.")
    headers = _normalize_headers(rows[0] if rows else [])
    body_rows = rows[1:] if rows else []
    normalized_rows = [_row_to_record(headers, row) for row in body_rows if _has_value(row)]
    safe_table_name = _safe_identifier(table_name or safe_source.stem, fallback="uploaded_data")
    sqlite_path = output_dir / f"{_safe_identifier(datasource_id, fallback='datasource')}.sqlite"
    _write_sqlite(sqlite_path, safe_table_name, headers, normalized_rows)
    return FileDataSourceImportResult(
        sqlite_path=str(sqlite_path),
        original_filename=Path(original_filename or safe_source.name).name,
        source_type=source_type,
        file_size=safe_source.stat().st_size,
        table_name=safe_table_name,
        uploaded_at=utc_now().isoformat(),
        row_count=len(normalized_rows),
        columns=headers,
    )


def validate_supported_file_path(path: Path, *, max_bytes: int | None = None) -> Path:
    """校验文件路径、扩展名和大小，避免读取明显敏感或不支持的文件。"""

    if ".." in path.parts:
        raise ValueError("Path traversal is not allowed for file datasources.")
    resolved = path.expanduser().resolve()
    if not resolved.is_file():
        raise ValueError(f"File datasource path does not exist or is not a file: {path}")
    lower_name = resolved.name.lower()
    lower_parts = {part.lower() for part in resolved.parts}
    if lower_name in SENSITIVE_FILE_NAMES or lower_name.startswith(".env"):
        raise ValueError("Sensitive environment files cannot be registered as datasources.")
    if lower_parts & SENSITIVE_PATH_PARTS:
        raise ValueError(
            "Sensitive system or credential paths cannot be registered as datasources."
        )
    if resolved.suffix.lower() not in ALLOWED_FILE_SUFFIXES:
        raise ValueError(
            "Unsupported file datasource type. Allowed extensions: csv, xlsx, parquet."
        )
    if max_bytes is not None and resolved.stat().st_size > max_bytes:
        raise ValueError(f"File exceeds upload limit of {max_bytes} bytes.")
    return resolved


def file_kind_for_path(path: Path) -> str:
    """根据文件扩展名返回 registry 使用的数据源 kind。"""

    suffix = path.suffix.lower()
    if suffix == ".csv":
        return "file_csv"
    if suffix == ".xlsx":
        return "file_excel"
    if suffix == ".parquet":
        return "file_parquet"
    raise ValueError("Unsupported file datasource type. Allowed extensions: csv, xlsx, parquet.")


def _read_rows(path: Path) -> list[list[Any]]:
    """按文件类型读取二维行数据，第一行必须作为表头。"""

    try:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return _read_csv_rows(path)
        if suffix == ".xlsx":
            return _read_excel_rows(path)
        if suffix == ".parquet":
            return _read_parquet_rows(path)
        raise ValueError(f"Unsupported file datasource extension: {suffix}")
    except RuntimeError:
        raise
    except Exception as exc:
        raise ValueError(f"Failed to parse file datasource: {exc}") from exc


def _read_csv_rows(path: Path) -> list[list[Any]]:
    """读取 UTF-8/UTF-8 BOM CSV 文件。"""

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [list(row) for row in csv.reader(handle)]


def _read_excel_rows(path: Path) -> list[list[Any]]:
    """读取 xlsx 第一个工作表，不执行公式，只读取缓存值。"""

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _read_parquet_rows(path: Path) -> list[list[Any]]:
    """读取 Parquet 文件；缺少 pyarrow 时返回清晰错误。"""

    try:
        import pyarrow.parquet as pq
    except ImportError as exc:
        raise RuntimeError(
            "Parquet datasource support requires optional dependency pyarrow."
        ) from exc
    table = pq.read_table(path)
    headers = [str(name) for name in table.column_names]
    return [headers, *[[record.get(header) for header in headers] for record in table.to_pylist()]]


def _normalize_headers(raw_headers: Iterable[Any]) -> list[str]:
    """将文件表头清洗为稳定 SQL 标识符，并处理重复列名。"""

    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, raw_header in enumerate(raw_headers, start=1):
        base = _safe_identifier(str(raw_header or ""), fallback=f"column_{index}")
        count = seen.get(base, 0)
        seen[base] = count + 1
        headers.append(base if count == 0 else f"{base}_{count + 1}")
    if not headers:
        raise ValueError("File datasource must contain a header row.")
    return headers


def _row_to_record(headers: list[str], row: list[Any]) -> dict[str, Any]:
    """把一行文件数据转换为 SQLite 插入记录。"""

    return {
        header: _normalize_cell(row[index] if index < len(row) else None)
        for index, header in enumerate(headers)
    }


def _has_value(row: list[Any]) -> bool:
    """过滤完全为空的行。"""

    return any(value not in (None, "") for value in row)


def _normalize_cell(value: Any) -> Any:
    """将日期等 Python 对象转换成 SQLite 可持久化值。"""

    if isinstance(value, datetime | date):
        return value.isoformat()
    if value == "":
        return None
    return value


def _write_sqlite(
    sqlite_path: Path,
    table_name: str,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    """创建受控 SQLite 表并批量写入文件数据。"""

    column_types = _infer_column_types(headers, rows)
    quoted_table = _quote_identifier(table_name)
    column_defs = ", ".join(
        f"{_quote_identifier(header)} {column_types[header]}" for header in headers
    )
    placeholders = ", ".join("?" for _ in headers)
    quoted_columns = ", ".join(_quote_identifier(header) for header in headers)
    sqlite_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(sqlite_path) as connection:
        connection.execute(f"DROP TABLE IF EXISTS {quoted_table}")
        connection.execute(f"CREATE TABLE {quoted_table} ({column_defs})")
        if rows:
            connection.executemany(
                f"INSERT INTO {quoted_table} ({quoted_columns}) VALUES ({placeholders})",
                [
                    [
                        _coerce_for_sqlite(row.get(header), column_types[header])
                        for header in headers
                    ]
                    for row in rows
                ],
            )


def _infer_column_types(headers: list[str], rows: list[dict[str, Any]]) -> dict[str, str]:
    """按列值推断 SQLite 类型，保障 SUM 等规则 SQL 能处理数字列。"""

    return {header: _infer_column_type([row.get(header) for row in rows]) for header in headers}


def _infer_column_type(values: list[Any]) -> str:
    """推断 INTEGER/REAL/TEXT 三类 SQLite 类型。"""

    non_null_values = [value for value in values if value not in (None, "")]
    if non_null_values and all(_is_int_like(value) for value in non_null_values):
        return "INTEGER"
    if non_null_values and all(_is_float_like(value) for value in non_null_values):
        return "REAL"
    return "TEXT"


def _coerce_for_sqlite(value: Any, column_type: str) -> Any:
    """按 SQLite 类型转换单元格值。"""

    if value in (None, ""):
        return None
    if column_type == "INTEGER":
        return int(float(value))
    if column_type == "REAL":
        return float(value)
    return str(value)


def _is_int_like(value: Any) -> bool:
    """判断值是否可安全视作整数。"""

    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        return False
    return numeric_value.is_integer()


def _is_float_like(value: Any) -> bool:
    """判断值是否可安全视作数字。"""

    try:
        float(value)
    except (TypeError, ValueError):
        return False
    return True


def _safe_identifier(value: str, *, fallback: str) -> str:
    """生成 SQL 友好的标识符，避免路径片段或特殊字符进入表/列名。"""

    normalized = IDENTIFIER_PATTERN.sub("_", value.strip().lower()).strip("_")
    if not normalized:
        normalized = fallback
    if normalized[0].isdigit():
        normalized = f"_{normalized}"
    return normalized[:64]


def _quote_identifier(value: str) -> str:
    """安全引用已清洗的 SQLite 标识符。"""

    return f'"{value.replace(chr(34), chr(34) * 2)}"'


__all__ = [
    "ALLOWED_FILE_SUFFIXES",
    "FileDataSourceImportResult",
    "file_kind_for_path",
    "import_file_to_sqlite",
    "validate_supported_file_path",
]
