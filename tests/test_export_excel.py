"""Tests for real Excel export artifacts."""

from io import BytesIO

from openpyxl import load_workbook

from persistence import InMemoryArtifactStore
from schemas import ChartSpec, ChartType, Insight, ReportFormat
from schemas.analysis_package import AnalysisPackage
from schemas.query_result import QueryColumn, QueryResult
from schemas.report import ReportOutline, ReportOutlineSection
from tools.export_tools import EXCEL_MIME_TYPE, export_excel


def test_export_excel_creates_readable_xlsx_artifact() -> None:
    """Excel export should persist a readable XLSX file in the artifact store."""

    artifact_store = InMemoryArtifactStore()
    package = _analysis_package()
    outline = _outline(package)

    result = export_excel(outline, artifact_store, package)

    assert result.report_format is ReportFormat.EXCEL
    assert result.status == "created"
    record = artifact_store.get_artifact(result.artifact_ref)
    assert record is not None
    assert isinstance(record.content, bytes)
    assert record.metadata["mime_type"] == EXCEL_MIME_TYPE
    assert record.metadata["report_type"] == "excel"
    assert record.metadata["source_analysis_package_id"] == package.package_id
    assert record.metadata["placeholder"] is False

    workbook = load_workbook(BytesIO(record.content))
    assert workbook.sheetnames == ["Summary", "Query Result", "Outline"]
    assert workbook["Summary"]["A1"].value == "Title"
    assert workbook["Summary"]["B1"].value == outline.title
    assert workbook["Query Result"]["A1"].value == "month"
    assert workbook["Query Result"]["B1"].value == "total_revenue"
    assert workbook["Query Result"]["A2"].value == "2026-01"
    assert workbook["Query Result"]["B3"].value == 210


def test_export_excel_metadata_is_readable_without_content() -> None:
    """Artifact metadata should expose mime type without reading XLSX bytes."""

    artifact_store = InMemoryArtifactStore()
    package = _analysis_package()
    result = export_excel(_outline(package), artifact_store, package)
    artifact_id = result.artifact_ref.rsplit(":", maxsplit=1)[-1]

    metadata = artifact_store.get_artifact_metadata(artifact_id)

    assert metadata is not None
    assert metadata.artifact_ref == result.artifact_ref
    assert metadata.mime_type == EXCEL_MIME_TYPE
    assert metadata.content_type == "bytes"


def _analysis_package() -> AnalysisPackage:
    """Create an analysis package with a tabular result."""

    return AnalysisPackage(
        question="Show monthly revenue trend",
        sql_result=QueryResult(
            sql="SELECT month, SUM(revenue) AS total_revenue FROM orders GROUP BY month",
            columns=[
                QueryColumn(name="month", data_type="text"),
                QueryColumn(name="total_revenue", data_type="real"),
            ],
            rows=[
                {"month": "2026-01", "total_revenue": 100.0},
                {"month": "2026-02", "total_revenue": 210.0},
            ],
            row_count=2,
        ),
        chart_spec=ChartSpec(
            chart_type=ChartType.LINE,
            title="Monthly revenue",
            artifact_ref="artifact:chart-1",
        ),
        insights=[
            Insight(
                title="Revenue trend",
                summary="Revenue increased across the period.",
            )
        ],
    )


def _outline(package: AnalysisPackage) -> ReportOutline:
    """Create an Excel outline tied to a package."""

    return ReportOutline(
        report_format=ReportFormat.EXCEL,
        title="Revenue workbook",
        sections=[ReportOutlineSection(title="Data", points=["Include query rows."])],
        source_package_id=package.package_id,
    )
